"""
lre.py — Motor LRE (Libro de Remuneraciones Electrónico)
=========================================================
Genera y exporta el archivo CSV normativo de 147 columnas
exigido por la Dirección del Trabajo (DT) de Chile.

Arquitectura:
  1. generate_lre → Toma liquidaciones ya calculadas, las cruza con
     contratos y employees, y persiste una "foto" inmutable en
     payroll_books + payroll_book_details.
  2. export_lre  → Lee esa foto y genera el CSV con los 147 campos DT.
  3. list_lre    → Lista los libros generados (sin caché, para evitar IDs obsoletos).
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from typing import Any
import io
import csv
import logging
from core.database import get_supabase
from core.utils.shared_utils import clean_rut_simple as clean_rut
from core.auth import verify_token
from core.payroll_status import closed_liquidation_statuses
from calculators.national_params import get_tramo_asignacion
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("lre_engine")
router = APIRouter()

# ─── 147 Columnas DT (Chile 2024) ────────────────────────────────
DT_HEADERS = [
    "Rut trabajador(1101)", "Fecha inicio contrato(1102)", "Fecha término de contrato(1103)",
    "Causal término de contrato(1104)", "Región prestación de servicios(1105)", "Comuna prestación de servicios(1106)",
    "Tipo impuesto a la renta(1170)", "Técnico extranjero exención cot. previsionales(1146)", "Código tipo de jornada(1107)",
    "Persona con Discapacidad - Pensionado por Invalidez(1108)", "Pensionado por vejez(1109)", "AFP(1141)",
    "IPS (ExINP)(1142)", "FONASA - ISAPRE(1143)", "AFC(1151)", "CCAF(1110)", "Org. administrador ley 16.744(1152)",
    "Nro cargas familiares legales autorizadas(1111)", "Nro de cargas familiares maternales(1112)", "Nro de cargas familiares invalidez(1113)",
    "Tramo asignación familiar(1114)", "Rut org sindical 1(1171)", "Rut org sindical 2(1172)", "Rut org sindical 3(1173)",
    "Rut org sindical 4(1174)", "Rut org sindical 5(1175)", "Rut org sindical 6(1176)", "Rut org sindical 7(1177)",
    "Rut org sindical 8(1178)", "Rut org sindical 9(1179)", "Rut org sindical 10(1180)", "Nro días trabajados en el mes(1115)",
    "Nro días de licencia médica en el mes(1116)", "Nro días de vacaciones en el mes(1117)", "Subsidio trabajador joven(1118)",
    "Puesto Trabajo Pesado(1154)", "APVI(1155)", "APVC(1157)", "Indemnización a todo evento(1131)", "Tasa indemnización a todo evento(1132)",
    "Sueldo(2101)", "Sobresueldo(2102)", "Comisiones(2103)", "Semana corrida(2104)", "Participación(2105)",
    "Gratificación(2106)", "Recargo 30% día domingo(2107)", "Remun. variable pagada en vacaciones(2108)",
    "Remun. variable pagada en clausura(2109)", "Aguinaldo(2110)", "Bonos u otras remun. fijas mensuales(2111)",
    "Tratos(2112)", "Bonos u otras remun. variables mensuales o superiores a un mes(2113)", "Ejercicio opción no pactada en contrato(2114)",
    "Beneficios en especie constitutivos de remun(2115)", "Remuneraciones bimestrales(2116)", "Remuneraciones trimestrales(2117)",
    "Remuneraciones cuatrimestral(2118)", "Remuneraciones semestrales(2119)", "Remuneraciones anuales(2120)",
    "Participación anual(2121)", "Gratificación anual(2122)", "Otras remuneraciones superiores a un mes(2123)",
    "Pago por horas de trabajo sindical(2124)", "Sueldo empresarial (2161)", "Subsidio por incapacidad laboral por licencia médica(2201)",
    "Beca de estudio(2202)", "Gratificaciones de zona(2203)", "Otros ingresos no constitutivos de renta(2204)",
    "Colación(2301)", "Movilización(2302)", "Viáticos(2303)", "Asignación de pérdida de caja(2304)",
    "Asignación de desgaste herramienta(2305)", "Asignación familiar legal(2311)", "Gastos por causa del trabajo(2306)",
    "Gastos por cambio de residencia(2307)", "Sala cuna(2308)", "Asignación trabajo a distancia o teletrabajo(2309)",
    "Depósito convenido hasta UF 900(2347)", "Alojamiento por razones de trabajo(2310)", "Asignación de traslación(2312)",
    "Indemnización por feriado legal(2313)", "Indemnización años de servicio(2314)", "Indemnización sustitutiva del aviso previo(2315)",
    "Indemnización fuero maternal(2316)", "Pago indemnización a todo evento(2331)", "Indemnizaciones voluntarias tributables(2417)",
    "Indemnizaciones contractuales tributables(2418)", "Cotización obligatoria previsional (AFP o IPS)(3141)",
    "Cotización obligatoria salud 7%(3143)", "Cotización voluntaria para salud(3144)", "Cotización AFC - trabajador(3151)",
    "Cotizaciones técnico extranjero para seguridad social fuera de Chile(3146)", "Descuento depósito convenido hasta UF 900 anual(3147)",
    "Cotización APVi Mod A(3155)", "Cotización APVi Mod B hasta UF50(3156)", "Cotización APVc Mod A(3157)",
    "Cotización APVc Mod B hasta UF50(3158)", "Impuesto retenido por remuneraciones(3161)", "Impuesto retenido por indemnizaciones(3162)",
    "Mayor retención de impuestos solicitada por el trabajador(3163)", "Impuesto retenido por reliquidación remun. devengadas otros períodos(3164)",
    "Diferencia impuesto reliquidación remun. devengadas en este período(3165)", "Retención préstamo clase media 2020 (Ley 21.252) (3166)",
    "Rebaja zona extrema DL 889 (3167)", "Cuota sindical 1(3171)", "Cuota sindical 2(3172)", "Cuota sindical 3(3173)",
    "Cuota sindical 4(3174)", "Cuota sindical 5(3175)", "Cuota sindical 6(3176)", "Cuota sindical 7(3177)",
    "Cuota sindical 8(3178)", "Cuota sindical 9(3179)", "Cuota sindical 10(3180)", "Crédito social CCAF(3110)",
    "Cuota vivienda o educación(3181)", "Crédito cooperativas de ahorro(3182)",
    "Otros descuentos autorizados y solicitados por el trabajador(3183)", "Cotización adicional trabajo pesado - trabajador(3154)",
    "Donaciones culturales y de reconstrucción(3184)", "Otros descuentos(3185)", "Pensiones de alimentos(3186)",
    "Descuento mujer casada(3187)", "Descuentos por anticipos y préstamos(3188)", "AFC - Aporte empleador(4151)",
    "Aporte empleador seguro accidentes del trabajo y Ley SANNA(4152)", "Aporte empleador indemnización a todo evento(4131)",
    "Aporte adicional trabajo pesado - empleador(4154)", "Aporte empleador seguro invalidez y sobrevivencia(4155)",
    "APVC - Aporte Empleador(4157)", "Total haberes(5201)", "Total haberes imponibles y tributables(5210)",
    "Total haberes imponibles no tributables(5220)", "Total haberes no imponibles y no tributables(5230)",
    "Total haberes no imponibles y tributables(5240)", "Total descuentos(5301)", "Total descuentos impuestos a las remuneraciones(5361)",
    "Total descuentos impuestos por indemnizaciones(5362)", "Total descuentos por cotizaciones del trabajador(5341)",
    "Total otros descuentos(5302)", "Total aportes empleador(5410)", "Total líquido(5501)", "Total indemnizaciones(5502)",
    "Total indemnizaciones tributables(5564)", "Total indemnizaciones no tributables(5565)"
]

# ─── Diccionarios institucionales DT ─────────────────────────────
AFP_DT_MAP = {
    'CAPITAL': '1', 'CUPRUM': '3', 'HABITAT': '5', 'MODELO': '34',
    'PLANVITAL': '29', 'PROVIDA': '8', 'UNO': '35'
}
SALUD_DT_MAP = {
    'FONASA': '7', 'BANMEDICA': '1', 'CONSALUD': '2', 'CRUZBLANCA': '3',
    'VIDATRES': '5', 'COLMENA': '6', 'NUEVA_MASVIDA': '30'
}
MUTUAL_DT_MAP = {'ACHS': '1', 'MUTUAL_CCHC': '2', 'MUSEG': '2', 'IST': '3', 'ISL': '4'}
CCAF_DT_MAP = {
    'LOS_ANDES': '101', 'LA_ARAUCANA': '102', '18_SEPTIEMBRE': '103',
    'LOS_HEROES': '107', '': '0', 'NINGUNA': '0'
}


# ─── Utilidades de Conversión Segura ─────────────────────────────
def safe_float(val: Any) -> float:
    try:
        if val is None or val == "": return 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def safe_int(val: Any) -> int:
    try:
        if val is None or val == "": return 0
        return int(round(float(val)))
    except (ValueError, TypeError):
        return 0


# ─── Modelo de Request ───────────────────────────────────────────
class LREGenerateRequest(BaseModel):
    organization_id: str
    periodo: str      # "YYYY-MM" o "YYYY-MM-DD"
    company_name: str
    company_rut: str


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ENDPOINT 1: LISTAR LIBROS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@router.get("/list")
async def list_lre(organization_id: str, current_user: dict = Depends(verify_token)):
    """Lista libros LRE. Consulta directa a DB (sin caché) para evitar IDs obsoletos."""
    db = get_supabase()
    try:
        res = db.table("payroll_books") \
            .select("*") \
            .eq("organization_id", organization_id) \
            .order("periodo", desc=True) \
            .execute()
        return res.data
    except Exception as e:
        logger.error(f"Error listando LRE: {e}")
        raise HTTPException(status_code=500, detail="Error al listar libros LRE")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ENDPOINT 2: GENERAR LIBRO
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@router.post("/generate")
async def generate_lre(req: LREGenerateRequest, current_user: dict = Depends(verify_token)):
    db = get_supabase()
    org_id = req.organization_id
    # Normalizar periodo: siempre YYYY-MM-DD para la DB
    periodo_std = f"{req.periodo}-01" if len(req.periodo) == 7 else req.periodo
    
    try:
        # 1. Recuperar liquidaciones con datos de empleados
        liq_res = db.table("liquidations") \
            .select("*, employees(*)") \
            .eq("organization_id", org_id) \
            .eq("periodo", periodo_std) \
            .in_("status", closed_liquidation_statuses()) \
            .execute()
        liqs = liq_res.data
        if not liqs:
            raise HTTPException(
                status_code=404,
                detail="No hay liquidaciones aprobadas o cerradas para generar el LRE de este periodo."
            )

        # 2. Calcular totales del libro
        t_haberes, t_descuentos, t_liquido = 0, 0, 0
        for l in liqs:
            t_haberes += safe_int(l.get("total_haberes_brutos", 0))
            t_descuentos += safe_int(l.get("total_descuentos", 0))
            t_liquido += safe_int(l.get("sueldo_liquido", 0))

        # 3. Datos cruzados: contratos y settings
        contracts_res = db.table("employment_contracts") \
            .select("*") \
            .eq("organization_id", org_id) \
            .execute()
        contracts_map = {c["employee_id"]: c for c in contracts_res.data}
        
        settings_res = db.table("organization_payroll_settings") \
            .select("*") \
            .eq("organization_id", org_id) \
            .execute()
        settings = settings_res.data[0] if settings_res.data else {}

        # 4. Limpiar libros previos del mismo periodo (regeneración limpia)
        prev_books = db.table("payroll_books") \
            .select("id") \
            .eq("organization_id", org_id) \
            .eq("periodo", periodo_std) \
            .execute()
        if prev_books.data:
            b_ids = [b["id"] for b in prev_books.data]
            db.table("payroll_book_details").delete().in_("payroll_book_id", b_ids).execute()
            db.table("payroll_books").delete().in_("id", b_ids).execute()

        seq_res = db.table("payroll_books") \
            .select("book_number") \
            .eq("organization_id", org_id) \
            .order("book_number", desc=True) \
            .limit(1) \
            .execute()
        next_book_number = safe_int(seq_res.data[0].get("book_number")) + 1 if seq_res.data else 1

        # 5. Crear encabezado del libro
        new_book = db.table("payroll_books").insert({
            "organization_id": org_id,
            "periodo": periodo_std,
            "book_number": next_book_number,
            "company_name": req.company_name,
            "company_rut": req.company_rut,
            "status": "approved",
            "total_employees": len(liqs),
            "total_haberes": t_haberes,
            "total_descuentos": t_descuentos,
            "total_liquido": t_liquido
        }).execute()
        
        book_id = new_book.data[0]["id"]

        # 6. Persistir detalles por trabajador (foto inmutable del periodo)
        details = []
        for l in liqs:
            emp = l.get("employees") or {}
            contract = contracts_map.get(emp.get("id"), {})
            
            details.append({
                "payroll_book_id": book_id, 
                "employee_id": emp.get("id"), 
                "employee_rut": emp.get("rut"),
                "apellido_paterno": emp.get("apellido_paterno", ""), 
                "apellido_materno": emp.get("apellido_materno", ""),
                "nombres": emp.get("nombres", ""), 
                "cargo": contract.get("cargo") or emp.get("cargo", "GENERAL"),
                "area": contract.get("department") or emp.get("departamento", "GENERAL"),
                "dias_trabajados": safe_int(l.get("dias_trabajados", 30)),
                "fecha_inicio": contract.get("fecha_inicio") or emp.get("fecha_ingreso") or "2024-01-01",
                "fecha_termino": contract.get("fecha_termino_fijo") or None,
                "causal_termino": contract.get("causal_termino") or "0",
                "region_prestacion": settings.get("region", "12"), 
                "comuna_prestacion": settings.get("comuna", "12101"),
                # Haberes
                "sueldo_base": safe_int(l.get("sueldo_base", 0)), 
                "sobresueldo": safe_int(l.get("horas_extra_monto", 0)),
                "gratificacion_legal": safe_int(l.get("gratificacion", 0)), 
                "bono_extra": safe_int(l.get("bono_extra", 0)),
                "colacion": safe_int(l.get("asignacion_colacion", 0)),
                "movilizacion": safe_int(l.get("asignacion_movilizacion", 0)), 
                "asig_familiar": safe_int(l.get("asignacion_familiar", 0)),
                "total_haberes_imponibles": safe_int(l.get("base_imponible_afp", 0)),
                "total_haberes_brutos": safe_int(l.get("total_haberes_brutos", 0)),
                # Instituciones
                "afp_nom": (emp.get("afp") or "HABITAT").upper(),
                "salud_nom": (emp.get("prevision_salud") or "FONASA").upper(),
                # Descuentos
                "descuento_afp_total": safe_int(safe_float(l.get("afp", 0)) + safe_float(l.get("afp_comision", 0))),
                "descuento_salud": safe_int(l.get("salud", 0)), 
                "salud_voluntaria": safe_int(l.get("salud_voluntaria", 0)),
                "afc_trab": safe_int(l.get("afc_trabajador", 0)), 
                "afc_emp": safe_int(l.get("afc_empresa", 0)),
                "sis_emp": safe_int(l.get("sis_empresa", 0)), 
                "impuesto_unico": safe_int(l.get("impuesto_unico", 0)),
                "total_descuentos": safe_int(l.get("total_descuentos", 0)),
                "sueldo_liquido": safe_int(l.get("sueldo_liquido", 0)),
                # Campos DT adicionales
                "family_allowances": safe_int(emp.get("family_allowances", 0)),
                "afc_active": bool(emp.get("afc_active", True)),
                "tipo_contrato": emp.get("tipo_contrato", "indefinido"),
            })
        
        db.table("payroll_book_details").insert(details).execute()
        
        logger.info(f"✅ LRE generado: book_id={book_id} periodo={periodo_std} empleados={len(liqs)}")
        return {"success": True, "book_id": book_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Fallo generación LRE: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ENDPOINT 3: EXPORTAR CSV (147 COLUMNAS DT)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@router.get("/export/{book_id}")
async def export_lre(book_id: str, current_user: dict = Depends(verify_token)):
    db = get_supabase()

    try:
        # 1. Recuperar libro (patrón robusto: sin single/maybe_single)
        book_res = db.table("payroll_books").select("*").eq("id", book_id).execute()
        if not book_res.data:
            raise HTTPException(
                status_code=404, 
                detail="El libro solicitado no existe. Probablemente fue reemplazado al regenerar. Refresque la página."
            )
        book = book_res.data[0]
        
        # 2. Recuperar detalles por trabajador
        details = db.table("payroll_book_details") \
            .select("*") \
            .eq("payroll_book_id", book_id) \
            .execute().data or []
        
        # 3. Recuperar configuración organizacional (Mutual, CCAF)
        org_id = book.get("organization_id")
        settings_res = db.table("organization_payroll_settings") \
            .select("*") \
            .eq("organization_id", org_id) \
            .execute()
        settings = settings_res.data[0] if settings_res.data else {}

        # Cargar liquidaciones del periodo para rebajas y snapshots
        liq_res = db.table("liquidations") \
            .select("employee_id, calculation_snapshot, asignacion_familiar") \
            .eq("organization_id", org_id) \
            .eq("periodo", book.get("periodo")) \
            .execute()
        liq_map = {l["employee_id"]: l for l in liq_res.data}

        # Cargar empleados para sueldos base nominales. El tramo familiar se calcula
        # desde la renta imponible porque employees.tramo_asignacion no existe en el
        # esquema actual.
        emp_res = db.table("employees") \
            .select("id, family_allowances, sueldo_base, afp") \
            .eq("organization_id", org_id) \
            .execute()
        emp_map = {e["id"]: e for e in emp_res.data}

        # 4. Generar CSV
        output = io.StringIO()
        output.write('\uFEFF')  # BOM para Excel
        writer = csv.DictWriter(output, fieldnames=DT_HEADERS, delimiter=';')
        writer.writeheader()
        
        TASAS_AFP_2026 = {
            'CAPITAL': 11.44, 'CUPRUM': 11.44, 'HABITAT': 11.27, 'MODELO': 10.58,
            'PLANVITAL': 11.16, 'PROVIDA': 11.45, 'UNO': 10.60
        }
        
        for d in details:
            emp_id = d.get("employee_id")
            liq_ref = liq_map.get(emp_id, {})
            emp_ref = emp_map.get(emp_id, {})
            
            # Inicializar todos los campos en "0" (formato estricto DT)
            row: dict[str, Any] = {h: "0" for h in DT_HEADERS}
            
            dias_trab = safe_int(d.get("dias_trabajados", 30))
            es_subsidio_total = (dias_trab == 0)
            
            # ── SECCIÓN 1000: DATOS ADMINISTRATIVOS ──
            row["Rut trabajador(1101)"] = clean_rut(d.get("employee_rut"))
            row["Fecha inicio contrato(1102)"] = d.get("fecha_inicio") or "2024-01-01"
            row["Fecha término de contrato(1103)"] = d.get("fecha_termino") or "0"
            row["Causal término de contrato(1104)"] = d.get("causal_termino") or "0"
            row["Región prestación de servicios(1105)"] = d.get("region_prestacion", "12")
            row["Comuna prestación de servicios(1106)"] = d.get("comuna_prestacion", "12101")
            row["Tipo impuesto a la renta(1170)"] = "1"
            row["Código tipo de jornada(1107)"] = "1"
            
            # Instituciones previsionales
            row["AFP(1141)"] = AFP_DT_MAP.get(str(d.get("afp_nom", "")).upper(), "5")
            row["FONASA - ISAPRE(1143)"] = SALUD_DT_MAP.get(str(d.get("salud_nom", "")).upper(), "7")
            row["AFC(1151)"] = "1" if d.get("afc_active", True) else "0"
            
            # Entidades empresariales
            m_code = str(settings.get("mutual_code", "ACHS")).upper()
            c_code = str(settings.get("caja_compensacion_code", "")).upper()
            row["Org. administrador ley 16.744(1152)"] = MUTUAL_DT_MAP.get(m_code, "1")
            row["CCAF(1110)"] = CCAF_DT_MAP.get(c_code, "0")
            
            # Cargas y días
            row["Nro cargas familiares legales autorizadas(1111)"] = str(safe_int(d.get("family_allowances", 0)))
            row["Nro días trabajados en el mes(1115)"] = str(dias_trab)
            
            # Tramo asignación familiar (1114) — mismos topes que el monto (SSoT)
            fam_allowances = safe_int(d.get("family_allowances", 0))
            tramo = "D"
            if fam_allowances > 0:
                tramo = get_tramo_asignacion(safe_int(d.get("total_haberes_imponibles", 0)))
            row["Tramo asignación familiar(1114)"] = tramo

            if es_subsidio_total:
                # Subsidio nominal LRE
                sueldo_pactado = safe_int(emp_ref.get("sueldo_base", 0))
                renta_ref = max(sueldo_pactado, 539000)
                
                row["Sueldo(2101)"] = str(renta_ref)
                row["Nro días de licencia médica en el mes(1116)"] = "30"
                row["Nro días trabajados en el mes(1115)"] = "0"
                
                # Calcular previsión nominal
                afp_nom = str(d.get("afp_nom", "MODELO")).upper()
                tasa_afp = TASAS_AFP_2026.get(afp_nom, 10.58) / 100
                afp_monto_nom = int(round(renta_ref * tasa_afp))
                row["Cotización obligatoria previsional (AFP o IPS)(3141)"] = str(afp_monto_nom)
                
                salud_monto_nom = int(round(renta_ref * 0.07))
                row["Cotización obligatoria salud 7%(3143)"] = str(salud_monto_nom)
                
                afc_trab_nom = 0
                afc_emp_nom = 0
                if "HONORARIO" not in d.get("tipo_contrato", "").upper():
                    if "indefinido" in d.get("tipo_contrato", "").lower():
                        afc_trab_nom = int(round(renta_ref * 0.006))
                        afc_emp_nom = int(round(renta_ref * 0.024))
                    else:
                        afc_emp_nom = int(round(renta_ref * 0.03))
                row["Cotización AFC - trabajador(3151)"] = str(afc_trab_nom)
                row["AFC - Aporte empleador(4151)"] = str(afc_emp_nom)
                
                sis_monto_nom = int(round(renta_ref * 0.0162))
                row["Aporte empleador seguro invalidez y sobrevivencia(4155)"] = str(sis_monto_nom)
                
                tasa_mutual = float(settings.get("tasa_mutual", 0.93)) / 100
                mutual_monto_nom = int(round(renta_ref * tasa_mutual))
                row["Aporte empleador seguro accidentes del trabajo y Ley SANNA(4152)"] = str(mutual_monto_nom)
                
                # Totales nominales
                row["Total haberes(5201)"] = str(renta_ref)
                row["Total haberes imponibles y tributables(5210)"] = str(renta_ref)
                row["Total haberes no imponibles y no tributables(5230)"] = "0"
                
                cot_trabajador = afp_monto_nom + salud_monto_nom + afc_trab_nom
                row["Total descuentos(5301)"] = str(cot_trabajador)
                row["Total descuentos por cotizaciones del trabajador(5341)"] = str(cot_trabajador)
                
                aportes_emp = afc_emp_nom + mutual_monto_nom + sis_monto_nom
                row["Total aportes empleador(5410)"] = str(aportes_emp)
                row["Total líquido(5501)"] = "0"
            else:
                # ── SECCIÓN 2000: HABERES ──
                snapshot = liq_ref.get("calculation_snapshot") or {}
                semana_corrida = safe_int(snapshot.get("semana_corrida", 0))
                viatico = safe_int(snapshot.get("asignacion_viatico", 0))
                otros_no_imp = safe_int(snapshot.get("otros_haberes_no_imponibles", 0))

                row["Sueldo(2101)"] = str(safe_int(d.get("sueldo_base", 0)))
                row["Sobresueldo(2102)"] = str(safe_int(d.get("sobresueldo", 0)))
                row["Semana corrida(2104)"] = str(semana_corrida)
                row["Gratificación(2106)"] = str(safe_int(d.get("gratificacion_legal", 0)))
                row["Bonos u otras remun. fijas mensuales(2111)"] = str(safe_int(d.get("bono_extra", 0)))
                row["Colación(2301)"] = str(safe_int(d.get("colacion", 0)))
                row["Movilización(2302)"] = str(safe_int(d.get("movilizacion", 0)))
                row["Viáticos(2303)"] = str(viatico)
                row["Otros ingresos no constitutivos de renta(2204)"] = str(otros_no_imp)
                row["Asignación familiar legal(2311)"] = str(safe_int(d.get("asig_familiar", 0)))

                # ── SECCIÓN 3000: DESCUENTOS ──
                row["Cotización obligatoria previsional (AFP o IPS)(3141)"] = str(safe_int(d.get("descuento_afp_total", 0)))
                row["Cotización obligatoria salud 7%(3143)"] = str(safe_int(d.get("descuento_salud", 0)))
                row["Cotización voluntaria para salud(3144)"] = str(safe_int(d.get("salud_voluntaria", 0)))
                row["Cotización AFC - trabajador(3151)"] = str(safe_int(d.get("afc_trab", 0)))
                row["Impuesto retenido por remuneraciones(3161)"] = str(safe_int(d.get("impuesto_unico", 0)))

                # ── SECCIÓN 4000: APORTES EMPLEADOR ──
                row["AFC - Aporte empleador(4151)"] = str(safe_int(d.get("afc_emp", 0)))
                row["Aporte empleador seguro invalidez y sobrevivencia(4155)"] = str(safe_int(d.get("sis_emp", 0)))
                
                # Aporte Mutual real de la empresa
                tasa_mutual = float(settings.get("tasa_mutual", 0.93)) / 100
                imponible = safe_int(d.get("total_haberes_imponibles", 0))
                mutual_monto = int(round(imponible * tasa_mutual))
                row["Aporte empleador seguro accidentes del trabajo y Ley SANNA(4152)"] = str(mutual_monto)

                # ── SECCIÓN 5000: TOTALES DT (calculados) ──
                haberes_brutos = safe_int(d.get("total_haberes_brutos", 0))
                haberes_imponibles = imponible
                colacion = safe_int(d.get("colacion", 0))
                movilizacion = safe_int(d.get("movilizacion", 0))
                asig_fam = safe_int(d.get("asig_familiar", 0))
                no_imponible = colacion + movilizacion + asig_fam + viatico + otros_no_imp
                
                descuento_afp = safe_int(d.get("descuento_afp_total", 0))
                descuento_salud = safe_int(d.get("descuento_salud", 0))
                salud_vol = safe_int(d.get("salud_voluntaria", 0))
                afc_trab = safe_int(d.get("afc_trab", 0))
                cot_trabajador = descuento_afp + descuento_salud + salud_vol + afc_trab
                
                afc_emp = safe_int(d.get("afc_emp", 0))
                sis_emp = safe_int(d.get("sis_emp", 0))
                aportes_emp = afc_emp + mutual_monto + sis_emp

                row["Total haberes(5201)"] = str(haberes_brutos)
                row["Total haberes imponibles y tributables(5210)"] = str(haberes_imponibles)
                row["Total haberes no imponibles y no tributables(5230)"] = str(no_imponible)
                row["Total descuentos(5301)"] = str(safe_int(d.get("total_descuentos", 0)))
                row["Total descuentos impuestos a las remuneraciones(5361)"] = str(safe_int(d.get("impuesto_unico", 0)))
                row["Total descuentos por cotizaciones del trabajador(5341)"] = str(cot_trabajador)
                row["Total aportes empleador(5410)"] = str(aportes_emp)
                row["Total líquido(5501)"] = str(safe_int(d.get("sueldo_liquido", 0)))

            # Rebaja zona extrema DL 889 (Campo 3167)
            rebaja_extrema = safe_int(liq_ref.get("calculation_snapshot", {}).get("rebaja_zona_extrema", 0))
            row["Rebaja zona extrema DL 889 (3167)"] = str(rebaja_extrema)

            # Sello de integridad: garantizar que no queden None
            for header in DT_HEADERS:
                if row[header] is None:
                    row[header] = "0"

            writer.writerow(row)
            
        content = output.getvalue()
        periodo_label = book.get('periodo', 'unknown')
        logger.info(f"📄 LRE Exportado: {book_id} | {book.get('company_name')} | {periodo_label}")
        
        return Response(
            content=content, 
            media_type="text/csv", 
            headers={
                "Content-Disposition": f"attachment; filename=LRE_{periodo_label}.csv",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exportando LRE: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error de sistema al exportar LRE: {str(e)}")

@router.get("/export-excel/{book_id}")
async def export_lre_excel(book_id: str, current_user: dict = Depends(verify_token)):
    """Genera el Excel de control y resumen detallado del Libro de Remuneraciones."""
    db = get_supabase()

    try:
        # 1. Recuperar libro
        book_res = db.table("payroll_books").select("*").eq("id", book_id).execute()
        if not book_res.data:
            raise HTTPException(
                status_code=404, 
                detail="El libro solicitado no existe."
            )
        book = book_res.data[0]
        
        # 2. Recuperar detalles por trabajador
        details = db.table("payroll_book_details") \
            .select("*") \
            .eq("payroll_book_id", book_id) \
            .execute().data or []
        
        # 3. Crear Libro Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Remuneraciones"
        ws.freeze_panes = "G7" # Congelar columnas de identificación

        # Estilos Premium
        fnt_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        fnt_section = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        fnt_header = Font(name="Arial", size=9, bold=True, color="1F4E79")
        fnt_data = Font(name="Arial", size=9)
        fnt_total = Font(name="Arial", size=9, bold=True)
        
        fill_title = PatternFill("solid", fgColor="1F4E79")
        fill_sec_id = PatternFill("solid", fgColor="2F75B6")
        fill_sec_hab = PatternFill("solid", fgColor="375623")
        fill_sec_desc = PatternFill("solid", fgColor="843C0C")
        fill_sec_liq = PatternFill("solid", fgColor="7B3F00")
        
        fill_hdr_id = PatternFill("solid", fgColor="D9E1F2")
        fill_hdr_hab = PatternFill("solid", fgColor="E2EFDA")
        fill_hdr_desc = PatternFill("solid", fgColor="FCE4D6")
        fill_hdr_liq = PatternFill("solid", fgColor="FFF2CC")
        fill_tot = PatternFill("solid", fgColor="BDD7EE")
        
        aln_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
        aln_l = Alignment(horizontal="left", vertical="center")
        aln_r = Alignment(horizontal="right", vertical="center")
        
        thin = Side(style="thin", color="D3D3D3")
        brd = Border(left=thin, right=thin, top=thin, bottom=thin)
        FMT_MONEY = '$#,##0'
        
        # Título Principal (Fila 1)
        ws.merge_cells("A1:U1")
        t = ws["A1"]
        t.value = f"LIBRO DE REMUNERACIONES — {book.get('company_name', '').upper()}"
        t.font = fnt_title
        t.fill = fill_title
        t.alignment = aln_c
        ws.row_dimensions[1].height = 30
        
        # Datos de Empresa (Fila 2)
        ws.cell(row=2, column=1, value="RUT Empresa:").font = fnt_header
        ws.cell(row=2, column=2, value=book.get('company_rut', ''))
        ws.cell(row=2, column=4, value="Período:").font = fnt_header
        ws.cell(row=2, column=5, value=book.get('periodo', ''))
        ws.cell(row=2, column=7, value="Total Empleados:").font = fnt_header
        ws.cell(row=2, column=8, value=len(details))

        # Encabezados de Sección (Fila 4)
        ws.merge_cells("A4:F4")
        sec_id = ws["A4"]
        sec_id.value = "IDENTIFICACIÓN"
        sec_id.font = fnt_section
        sec_id.fill = fill_sec_id
        sec_id.alignment = aln_c
        
        ws.merge_cells("G4:N4")
        sec_hab = ws["G4"]
        sec_hab.value = "HABERES"
        sec_hab.font = fnt_section
        sec_hab.fill = fill_sec_hab
        sec_hab.alignment = aln_c
        
        ws.merge_cells("O4:T4")
        sec_desc = ws["O4"]
        sec_desc.value = "DESCUENTOS"
        sec_desc.font = fnt_section
        sec_desc.fill = fill_sec_desc
        sec_desc.alignment = aln_c
        
        ws.cell(row=4, column=21, value="LÍQUIDO").font = fnt_section
        ws.cell(row=4, column=21).fill = fill_sec_liq
        ws.cell(row=4, column=21).alignment = aln_c
        ws.row_dimensions[4].height = 20

        # Encabezados de Columna (Fila 5)
        HEADERS = [
            ("N°", fill_hdr_id),
            ("RUT Trabajador", fill_hdr_id),
            ("Nombre Completo", fill_hdr_id),
            ("Cargo", fill_hdr_id),
            ("Área / CC", fill_hdr_id),
            ("Días", fill_hdr_id),
            # Haberes
            ("Sueldo Base", fill_hdr_hab),
            ("Horas Extras", fill_hdr_hab),
            ("Gratificación", fill_hdr_hab),
            ("Bono Extra", fill_hdr_hab),
            ("Colación", fill_hdr_hab),
            ("Movilización", fill_hdr_hab),
            ("Asig. Familiar", fill_hdr_hab),
            ("Total Haberes", fill_hdr_hab),
            # Descuentos
            ("Cotiz. AFP", fill_hdr_desc),
            ("Cotiz. Salud", fill_hdr_desc),
            ("Salud Adic.", fill_hdr_desc),
            ("Seguro Cesantía", fill_hdr_desc),
            ("Impuesto Único", fill_hdr_desc),
            ("Total Descuentos", fill_hdr_desc),
            # Líquido
            ("Líquido a Pagar", fill_hdr_liq)
        ]

        for col, (h, fill) in enumerate(HEADERS, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = fnt_header
            cell.fill = fill
            cell.alignment = aln_c
            cell.border = brd
        ws.row_dimensions[5].height = 25

        # Escribir Filas de Datos (Filas 6+)
        start_row = 6
        for i, d in enumerate(details, 1):
            row = start_row + i - 1
            ws.row_dimensions[row].height = 20
            
            fullname = f"{d.get('apellido_paterno', '')} {d.get('apellido_materno', '')} {d.get('nombres', '')}".strip()
            
            ws.cell(row=row, column=1, value=i).alignment = aln_c
            ws.cell(row=row, column=2, value=d.get('employee_rut', '')).alignment = aln_c
            ws.cell(row=row, column=3, value=fullname).alignment = aln_l
            ws.cell(row=row, column=4, value=d.get('cargo', '')).alignment = aln_l
            ws.cell(row=row, column=5, value=d.get('area', '')).alignment = aln_l
            ws.cell(row=row, column=6, value=safe_int(d.get('dias_trabajados', 30))).alignment = aln_c
            
            # Haberes
            ws.cell(row=row, column=7, value=safe_int(d.get('sueldo_base', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=8, value=safe_int(d.get('sobresueldo', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=9, value=safe_int(d.get('gratificacion_legal', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=10, value=safe_int(d.get('bono_extra', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=11, value=safe_int(d.get('colacion', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=12, value=safe_int(d.get('movilizacion', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=13, value=safe_int(d.get('asig_familiar', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=14, value=safe_int(d.get('total_haberes_brutos', 0))).number_format = FMT_MONEY
            
            # Descuentos
            ws.cell(row=row, column=15, value=safe_int(d.get('descuento_afp_total', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=16, value=safe_int(d.get('descuento_salud', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=17, value=safe_int(d.get('salud_voluntaria', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=18, value=safe_int(d.get('afc_trab', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=19, value=safe_int(d.get('impuesto_unico', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=20, value=safe_int(d.get('total_descuentos', 0))).number_format = FMT_MONEY
            
            # Líquido
            ws.cell(row=row, column=21, value=safe_int(d.get('sueldo_liquido', 0))).number_format = FMT_MONEY
            ws.cell(row=row, column=21).fill = fill_hdr_liq
            ws.cell(row=row, column=21).font = fnt_total
            
            # Formato de celdas
            for col in range(1, 22):
                c = ws.cell(row=row, column=col)
                if col not in [3, 4, 5]:
                    c.alignment = aln_r if col >= 7 else aln_c
                if col != 21:
                    c.font = fnt_data
                c.border = brd

        # Fila de Totales
        tot_row = start_row + len(details)
        ws.row_dimensions[tot_row].height = 22
        
        ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=5)
        lbl_tot = ws.cell(row=tot_row, column=1, value="TOTALES GENERALES")
        lbl_tot.font = fnt_total
        lbl_tot.alignment = aln_c
        lbl_tot.fill = fill_tot
        lbl_tot.border = brd
        
        # Bordes para celdas combinadas de totales
        for col in range(2, 6):
            ws.cell(row=tot_row, column=col).border = brd
            ws.cell(row=tot_row, column=col).fill = fill_tot

        # Días totales
        c_dias = ws.cell(row=tot_row, column=6, value=f"=SUM(F{start_row}:F{tot_row-1})")
        c_dias.font = fnt_total
        c_dias.alignment = aln_c
        c_dias.fill = fill_tot
        c_dias.border = brd
        
        # Fórmulas de Suma para valores monetarios
        for col in range(7, 22):
            col_let = get_column_letter(col)
            c = ws.cell(row=tot_row, column=col, value=f"=SUM({col_let}{start_row}:{col_let}{tot_row-1})")
            c.font = fnt_total
            c.alignment = aln_r
            c.number_format = FMT_MONEY
            c.fill = fill_tot
            c.border = brd

        # Ajuste Automático de Anchos de Columna
        for col in range(1, 22):
            col_letter = get_column_letter(col)
            max_len = 0
            # Examinar fila 5 (header) y filas de datos
            for r in range(5, tot_row + 1):
                val = ws.cell(row=r, column=col).value
                if val:
                    val_str = str(val)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Guardar en memoria y retornar
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        filename = f"LibroRemuneraciones_{book.get('company_name', 'Control')}_{book.get('periodo', 'unknown')}.xlsx"
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generando Excel LRE: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")
