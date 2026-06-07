from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime
from lxml import etree
import io
import re
from core.database import get_supabase
from core.auth import verify_token, verify_org_role
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter()

# Códigos AFP según SII (Formulario 1887)
CODIGOS_AFP_SII = {
    "CAPITAL":   "08",
    "CUPRUM":    "02",
    "HABITAT":   "05",
    "MODELO":    "33",
    "PLANVITAL": "29",
    "PROVIDA":   "07",
    "UNO":       "34",
    "SIP":       "99",
    "IPS":       "99",
}

def clean_rut(rut: str) -> str:
    """Elimina puntos y guion de un RUT."""
    if not rut:
        return ""
    return re.sub(r'[^0-9kK]', '', rut)

def format_rut(rut: str) -> str:
    """Formatea un RUT como XX.XXX.XXX-X."""
    cleaned = clean_rut(rut)
    if len(cleaned) < 2:
        return cleaned
    body = cleaned[:-1]
    dv = cleaned[-1].upper()
    
    formatted_body = ""
    for i, char in enumerate(reversed(body)):
        if i > 0 and i % 3 == 0:
            formatted_body = "." + formatted_body
        formatted_body = char + formatted_body
    return f"{formatted_body}-{dv}"

def calcular_datos_anuales(liquidations: list) -> list:
    """Agrupa liquidaciones del año por empleado y calcula acumulados anuales."""
    datos_por_trab = {}
    for liq in liquidations:
        emp = liq.get("employees")
        if not emp:
            continue
        emp_id = emp.get("id")
        if emp_id not in datos_por_trab:
            datos_por_trab[emp_id] = {
                "employee": emp,
                "meses": 0,
                "renta_imponible": 0,
                "renta_no_imponible": 0,
                "impuesto": 0,
                "cotiz_afp": 0,
                "cotiz_salud": 0,
                "cotiz_salud_adic": 0,
                "cotiz_cesantia": 0,
            }
        
        d = datos_por_trab[emp_id]
        d["meses"] += 1
        
        snapshot = liq.get("calculation_snapshot") or {}
        semana_corrida = int(snapshot.get("semana_corrida") or 0)
        otros_hab_imp = int(snapshot.get("otros_haberes_imponibles") or 0)
        
        if "sueldo_base" not in liq:
            colacion = int(liq.get("asignacion_colacion") or liq.get("bono_colacion") or 0)
            movilizacion = int(liq.get("asignacion_movilizacion") or liq.get("bono_movilizacion") or 0)
            asig_familiar = int(liq.get("asignacion_familiar") or 0)
            viaticos = int(snapshot.get("asignacion_viatico") or 0)
            otros_no_imp = int(snapshot.get("otros_haberes_no_imponibles") or 0)
            r_imp = int(liq.get("total_haberes_brutos", 0) - colacion - movilizacion - asig_familiar - viaticos - otros_no_imp)
        else:
            r_imp = int(
                liq.get("sueldo_base", 0) +
                liq.get("gratificacion", 0) +
                liq.get("horas_extra_monto", 0) +
                liq.get("bono_extra", 0) +
                semana_corrida +
                otros_hab_imp
            )
        r_no_imp = int(liq.get("total_haberes_brutos", 0) - r_imp)
        
        d["renta_imponible"] += r_imp
        d["renta_no_imponible"] += r_no_imp
        
        # Descuentos
        d["cotiz_afp"] += int(liq.get("afp", 0) + liq.get("afp_comision", 0))
        d["cotiz_salud"] += int(liq.get("salud", 0))
        d["cotiz_salud_adic"] += int(liq.get("salud_voluntaria", 0))
        d["cotiz_cesantia"] += int(liq.get("afc_trabajador", 0))
        d["impuesto"] += int(liq.get("impuesto_unico", 0))

    return sorted(datos_por_trab.values(), key=lambda x: x["employee"].get("apellido_paterno", ""))

@router.get("/export-xml/{organization_id}")
async def export_dj1887_xml(organization_id: str, anio: int, current_user: dict = Depends(verify_token)):
    """Genera el XML oficial de la DJ1887 compatible con carga masiva del SII."""
    await verify_org_role(organization_id, auth=current_user)
    db = get_supabase()

    # 1. Obtener organización
    org_res = db.table("organizations").select("*").eq("id", organization_id).single().execute()
    org = org_res.data
    if not org:
        raise HTTPException(status_code=404, detail="Organización no encontrada")

    # 2. Obtener liquidaciones
    date_start = f"{anio}-01-01"
    date_end = f"{anio}-12-31"
    liq_res = db.table("liquidations") \
        .select("*, employees(*)") \
        .eq("organization_id", organization_id) \
        .gte("periodo", date_start) \
        .lte("periodo", date_end) \
        .execute()
    
    liquidations = liq_res.data or []
    if not liquidations:
        raise HTTPException(status_code=404, detail=f"No hay remuneraciones para el año {anio}")

    registros = calcular_datos_anuales(liquidations)

    # 3. Construir XML
    root = etree.Element("DJ1887")
    root.set("version", "1.0")
    
    car = etree.SubElement(root, "Caratula")
    etree.SubElement(car, "TipoFormulario").text = "1887"
    etree.SubElement(car, "RutEmisor").text = clean_rut(org.get("rut_empresa", ""))
    etree.SubElement(car, "RazonSocial").text = org.get("nombre", "").upper()
    etree.SubElement(car, "Periodo").text = str(anio)
    etree.SubElement(car, "NroRegistros").text = str(len(registros))
    etree.SubElement(car, "FechaGeneracion").text = datetime.now().strftime("%Y-%m-%d")

    tot_imp = sum(r["renta_imponible"] for r in registros)
    tot_nimp = sum(r["renta_no_imponible"] for r in registros)
    tot_ret = sum(r["impuesto"] for r in registros)
    etree.SubElement(car, "TotalRentaImponible").text = str(tot_imp)
    etree.SubElement(car, "TotalRentaNoImponible").text = str(tot_nimp)
    etree.SubElement(car, "TotalImpuestoRetenido").text = str(tot_ret)

    det = etree.SubElement(root, "Detalle")
    for reg in registros:
        emp = reg["employee"]
        r = etree.SubElement(det, "Registro")
        etree.SubElement(r, "RutTrabajador").text = clean_rut(emp.get("rut", ""))
        nombre = f"{emp.get('apellido_paterno', '')} {emp.get('apellido_materno', '')} {emp.get('nombres', '')}".strip().upper()
        etree.SubElement(r, "NombreTrabajador").text = nombre
        etree.SubElement(r, "MesesTrabajados").text = str(reg["meses"])

        etree.SubElement(r, "RentaImponible").text = str(reg["renta_imponible"])
        etree.SubElement(r, "RentaNoImponible").text = str(reg["renta_no_imponible"])
        etree.SubElement(r, "ImpuestoUnicoRetenido").text = str(reg["impuesto"])

        # AFP
        afp_upper = (emp.get("afp") or "MODELO").upper()
        etree.SubElement(r, "AFP").text = afp_upper
        etree.SubElement(r, "CodigoAFP").text = CODIGOS_AFP_SII.get(afp_upper, "33")
        etree.SubElement(r, "CotizacionAFP").text = str(reg["cotiz_afp"])

        # Salud
        salud_upper = (emp.get("prevision_salud") or "FONASA").upper()
        etree.SubElement(r, "SistemaSalud").text = salud_upper
        etree.SubElement(r, "CotizacionSalud").text = str(reg["cotiz_salud"])
        if reg["cotiz_salud_adic"] > 0:
            etree.SubElement(r, "CotizacionSaludAdicional").text = str(reg["cotiz_salud_adic"])

        # Cesantía
        etree.SubElement(r, "CotizacionCesantia").text = str(reg["cotiz_cesantia"])

    xml_bytes = etree.tostring(root, pretty_print=True, xml_declaration=True, encoding="UTF-8")
    
    filename = f"DJ1887_{clean_rut(org.get('rut_empresa', ''))}_{anio}.xml"
    return StreamingResponse(
        io.BytesIO(xml_bytes),
        media_type="application/xml",
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@router.get("/export-excel/{organization_id}")
async def export_dj1887_excel(organization_id: str, anio: int, current_user: dict = Depends(verify_token)):
    """Genera el Excel resumen de control de la DJ1887."""
    await verify_org_role(organization_id, auth=current_user)
    db = get_supabase()

    # 1. Obtener organización
    org_res = db.table("organizations").select("*").eq("id", organization_id).single().execute()
    org = org_res.data
    if not org:
        raise HTTPException(status_code=404, detail="Organización no encontrada")

    # 2. Obtener liquidaciones
    date_start = f"{anio}-01-01"
    date_end = f"{anio}-12-31"
    liq_res = db.table("liquidations") \
        .select("*, employees(*)") \
        .eq("organization_id", organization_id) \
        .gte("periodo", date_start) \
        .lte("periodo", date_end) \
        .execute()
    
    liquidations = liq_res.data or []
    if not liquidations:
        raise HTTPException(status_code=404, detail=f"No hay remuneraciones para el año {anio}")

    registros = calcular_datos_anuales(liquidations)

    # 3. Crear Libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"DJ1887_{anio}"

    # Estilos Premium
    fnt_title  = Font(bold=True, size=12)
    fnt_header = Font(bold=True, size=10, color="FFFFFF")
    fill_hdr   = PatternFill("solid", fgColor="1F4E79")
    fill_alt   = PatternFill("solid", fgColor="F2F4F7")
    fill_tot   = PatternFill("solid", fgColor="D9E1F2")
    aln_c      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    aln_r      = Alignment(horizontal="right",  vertical="center")
    thin       = Side(style="thin", color="D3D3D3")
    brd        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Encabezados
    ws.merge_cells("A1:N1")
    t = ws["A1"]
    t.value = f"DECLARACIÓN JURADA 1887 — REMUNERACIONES AÑO COMERCIAL {anio}"
    t.font = fnt_title
    t.alignment = aln_c
    ws.row_dimensions[1].height = 25

    ws.merge_cells("A2:N2")
    e = ws["A2"]
    e.value = f"Empresa: {org.get('nombre')}   RUT: {format_rut(org.get('rut_empresa', ''))}"
    e.alignment = aln_c
    ws.row_dimensions[2].height = 20

    HEADERS = [
        "N°", "RUT Trabajador", "Nombre Trabajador",
        "AFP", "Meses\nTrab.",
        "Renta\nImponible", "Renta\nNo Imponible",
        "Cotiz.\nAFP", "Cotiz.\nSalud", "Cotiz. Salud\nAdicional",
        "Cotiz.\nCesantía", "Impuesto Único\nRetenido",
        "Total\nRetenciones", "Líquido\nAprox."
    ]

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = fnt_header
        c.fill = fill_hdr
        c.alignment = aln_c
        c.border = brd
    ws.row_dimensions[4].height = 35

    # Llenar Datos
    for i, reg in enumerate(registros, 1):
        row = i + 4
        emp = reg["employee"]
        fill = fill_alt if i % 2 == 0 else None

        def write_cell(col, val, aln=None, is_money=False):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = brd
            if fill:
                cell.fill = fill
            if aln:
                cell.alignment = aln
            if is_money:
                cell.number_format = "$#,##0"
                cell.alignment = aln_r
            return cell

        write_cell(1, i, aln_c)
        write_cell(2, format_rut(emp.get("rut", "")), aln_c)
        write_cell(3, f"{emp.get('apellido_paterno', '')} {emp.get('apellido_materno', '')} {emp.get('nombres', '')}".strip().upper())
        write_cell(4, (emp.get("afp") or "MODELO").upper(), aln_c)
        write_cell(5, reg["meses"], aln_c)
        
        write_cell(6, reg["renta_imponible"], is_money=True)
        write_cell(7, reg["renta_no_imponible"], is_money=True)
        write_cell(8, reg["cotiz_afp"], is_money=True)
        write_cell(9, reg["cotiz_salud"], is_money=True)
        write_cell(10, reg["cotiz_salud_adic"], is_money=True)
        write_cell(11, reg["cotiz_cesantia"], is_money=True)
        write_cell(12, reg["impuesto"], is_money=True)

        tot_ret = reg["cotiz_afp"] + reg["cotiz_salud"] + reg["cotiz_salud_adic"] + reg["cotiz_cesantia"] + reg["impuesto"]
        liq_aprox = reg["renta_imponible"] + reg["renta_no_imponible"] - tot_ret
        
        write_cell(13, tot_ret, is_money=True)
        write_cell(14, liq_aprox, is_money=True)
        ws.row_dimensions[row].height = 22

    # Fila de Totales
    tot_row = len(registros) + 5
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=5)
    tot_lbl = ws.cell(row=tot_row, column=1, value="TOTAL CONSOLIDADO")
    tot_lbl.font = Font(bold=True)
    tot_lbl.fill = fill_tot
    tot_lbl.alignment = aln_c
    tot_lbl.border = brd

    for col in range(6, 15):
        c = ws.cell(row=tot_row, column=col)
        c.font = Font(bold=True)
        c.fill = fill_tot
        c.border = brd
        col_letter = ws.cell(row=5, column=col).column_letter
        if col < 13:
            c.value = f"=SUM({col_letter}5:{col_letter}{tot_row-1})"
            c.number_format = "$#,##0"
            c.alignment = aln_r
        elif col == 13:
            c.value = f"=SUM(H{tot_row}+I{tot_row}+J{tot_row}+K{tot_row}+L{tot_row})"
            c.number_format = "$#,##0"
            c.alignment = aln_r
        elif col == 14:
            c.value = f"=SUM(F{tot_row}+G{tot_row}-M{tot_row})"
            c.number_format = "$#,##0"
            c.alignment = aln_r
    
    ws.row_dimensions[tot_row].height = 25

    # Configurar Anchos
    widths = [5, 15, 35, 12, 8, 15, 15, 14, 14, 15, 14, 15, 15, 15]
    for col, w in enumerate(widths, 1):
        col_letter = ws.cell(row=4, column=col).column_letter
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A5"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"DJ1887_{clean_rut(org.get('rut_empresa', ''))}_{anio}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )
